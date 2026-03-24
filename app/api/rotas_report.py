"""
Canal Educação v3.0 — Rotas de Relatórios de Aula (Class Reports).
CRUD completo com auto-preenchimento, validação relacional,
campos condicionais, Naming Engine, paginação server-side e exportação CSV.
"""

import csv
import io
from datetime import date
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_current_active_user, require_role
from app.core.database import get_db
from app.models.audit_log import AuditLog
from app.models.auxiliares import Disciplina, Professor, ProfessorDisciplina
from app.models.class_report import ClassReport
from app.models.turma import Turma
from app.models.user import User
from app.schemas.class_report import (
    ClassReportCancel,
    ClassReportCreate,
    ClassReportResponse,
    PaginatedReportsResponse,
)
from app.services.naming_engine import (
    gerar_nome_padronizado,
    verificar_sufixo_geminada,
)

router = APIRouter(
    prefix="/api/v1/reports",
    tags=["Relatórios de Aula"],
)


# ─── Validação de Campos Condicionais ────────────────────────────────


def _validar_campos_condicionais(data: ClassReportCreate) -> None:
    """
    Valida campos condicionais conforme spec 3.C/3.D/3.E:
    - Se interação='Outras' → interacao_outras_desc obrigatório
    - Se recurso contém 'Outro' → recursos_outro_desc obrigatório
    - Se problema_material='Outros' → problema_material_outro_desc obrigatório
    - Se teve_substituicao → professor_substituto_id obrigatório
    - Se teve_atraso → minutos_atraso e observacao_atraso obrigatórios
    """
    if data.interacao_professor_aluno == "Outras" and not data.interacao_outras_desc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Campo 'interacao_outras_desc' é obrigatório quando interação é 'Outras'.",
        )

    if data.tipo_recursos_utilizados and "Outro" in data.tipo_recursos_utilizados:
        if not data.recursos_outro_desc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Campo 'recursos_outro_desc' é obrigatório quando recurso 'Outro' está selecionado.",
            )

    if data.problema_material == "Outros" and not data.problema_material_outro_desc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Campo 'problema_material_outro_desc' é obrigatório quando problema é 'Outros'.",
        )

    if data.teve_substituicao and not data.professor_substituto_id:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Campo 'professor_substituto_id' é obrigatório quando há substituição.",
        )

    if data.teve_atraso:
        if not data.minutos_atraso:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Campo 'minutos_atraso' é obrigatório quando há atraso.",
            )
        if not data.observacao_atraso:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Campo 'observacao_atraso' é obrigatório quando há atraso.",
            )


# ─── Validação Relacional Professor ↔ Disciplina ────────────────────


def _validar_professor_disciplina(
    db: Session, professor_id: UUID, disciplina_id: UUID
) -> None:
    """
    Compliance Crítico: Verifica se o professor está vinculado à disciplina.
    Apenas professores cadastrados para aquela disciplina podem ser selecionados.
    """
    vinculo = (
        db.query(ProfessorDisciplina)
        .filter(
            ProfessorDisciplina.professor_id == professor_id,
            ProfessorDisciplina.disciplina_id == disciplina_id,
        )
        .first()
    )

    if not vinculo:
        professor = db.query(Professor).filter(Professor.id == professor_id).first()
        disciplina = (
            db.query(Disciplina).filter(Disciplina.id == disciplina_id).first()
        )
        prof_nome = professor.nome if professor else str(professor_id)
        disc_nome = disciplina.nome if disciplina else str(disciplina_id)
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Validação Relacional: Professor '{prof_nome}' não está "
                f"vinculado à disciplina '{disc_nome}'."
            ),
        )


# ─── Filtro Helper ───────────────────────────────────────────────────


def _apply_report_filters(query, **kwargs):
    """Aplica filtros ao query de relatórios (reutilizado em list e export)."""
    if kwargs.get("data_aula"):
        query = query.filter(ClassReport.data_aula == kwargs["data_aula"])
    if kwargs.get("data_inicio"):
        query = query.filter(ClassReport.data_aula >= kwargs["data_inicio"])
    if kwargs.get("data_fim"):
        query = query.filter(ClassReport.data_aula <= kwargs["data_fim"])
    if kwargs.get("turma_id"):
        query = query.filter(ClassReport.turma_id == kwargs["turma_id"])
    if kwargs.get("professor_id"):
        query = query.filter(ClassReport.professor_id == kwargs["professor_id"])
    if kwargs.get("disciplina_id"):
        query = query.filter(ClassReport.disciplina_id == kwargs["disciplina_id"])
    if kwargs.get("status_filter"):
        query = query.filter(ClassReport.status == kwargs["status_filter"].upper())
    if kwargs.get("estudio"):
        query = query.filter(ClassReport.estudio == kwargs["estudio"])
    if kwargs.get("tipo_aula"):
        query = query.filter(ClassReport.tipo_aula == kwargs["tipo_aula"])
    # Boolean occurrence filters (§5.3 — critical for audit)
    if kwargs.get("com_atraso") is True:
        query = query.filter(ClassReport.teve_atraso == True)  # noqa: E712
    if kwargs.get("com_substituicao") is True:
        query = query.filter(ClassReport.teve_substituicao == True)  # noqa: E712
    if kwargs.get("com_problema"):
        query = query.filter(ClassReport.problema_material != "Não")
    return query


# ─── CRUD ─────────────────────────────────────────────────────────────


@router.post("/", response_model=ClassReportResponse, status_code=status.HTTP_201_CREATED)
def criar_relatorio(
    data: ClassReportCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Cria um novo relatório de aula (status RASCUNHO).
    Auto-preenche os metadados e valida campos condicionais.
    """
    # 1. Validação das entidades referenciadas
    turma = db.query(Turma).filter(Turma.id == data.turma_id).first()
    if not turma:
        raise HTTPException(status_code=404, detail="Turma não encontrada.")

    professor = db.query(Professor).filter(Professor.id == data.professor_id).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor não encontrado.")

    disciplina = (
        db.query(Disciplina).filter(Disciplina.id == data.disciplina_id).first()
    )
    if not disciplina:
        raise HTTPException(status_code=404, detail="Disciplina não encontrada.")

    # 2. Validação relacional Professor ↔ Disciplina
    _validar_professor_disciplina(db, data.professor_id, data.disciplina_id)

    # 3. Validação de campos condicionais
    _validar_campos_condicionais(data)

    # 4. Criação do relatório
    report = ClassReport(
        created_by=current_user.id,
        status="RASCUNHO",
        **data.model_dump(),
    )

    db.add(report)
    db.flush()
    if current_user.role == "admin":
        log = AuditLog(
            actor_id=current_user.id,
            action="CREATE_REPORT",
            details=f"Admin criou o relatório {report.id}"
        )
        db.add(log)
    db.commit()
    db.refresh(report)

    return report


@router.patch("/{report_id}/finalizar", response_model=ClassReportResponse)
def finalizar_relatorio(
    report_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Finaliza o relatório: executa o Naming Engine e verifica aulas geminadas (P1/P2/P3).
    """
    report = db.query(ClassReport).filter(ClassReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado.")

    if report.status != "RASCUNHO":
        raise HTTPException(
            status_code=400,
            detail=f"Relatório com status '{report.status}' não pode ser finalizado.",
        )

    # Restrição RBAC (Ownership)
    if current_user.role == "assistente" and report.created_by != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Assistentes só podem finalizar relatórios criados por eles mesmos.",
        )

    # Buscar turma para dados do naming engine
    turma = db.query(Turma).filter(Turma.id == report.turma_id).first()
    disciplina = (
        db.query(Disciplina).filter(Disciplina.id == report.disciplina_id).first()
    )

    # Gerar nome padronizado via Naming Engine
    nome_gerado = gerar_nome_padronizado(
        nomenclatura_turma=turma.nomenclatura_padrao or turma.nome,
        disciplina=disciplina.nome if disciplina else "",
        data_aula=report.data_aula,
        conteudo=report.conteudo_ministrado,
    )

    # Verificação de aulas geminadas (P1/P2/P3)
    aula_existente = (
        db.query(ClassReport)
        .filter(
            ClassReport.turma_id == report.turma_id,
            ClassReport.disciplina_id == report.disciplina_id,
            ClassReport.data_aula == report.data_aula,
            ClassReport.status == "FINALIZADO",
            ClassReport.id != report.id,
        )
        .first()
    )

    if aula_existente:
        if not verificar_sufixo_geminada(report.conteudo_ministrado):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    "Aula geminada detectada! Já existe uma aula finalizada para "
                    f"esta turma/disciplina/data. Adicione 'P1', 'P2' ou 'P3' ao "
                    "conteúdo ministrado para distinguir as aulas."
                ),
            )
        report.conflito_geminada_resolvido = True

    # Finaliza
    report.nome_ficheiro_gerado = nome_gerado
    report.status = "FINALIZADO"

    db.commit()
    db.refresh(report)

    # Dispara task assíncrona de compliance no Google Drive (§6)
    try:
        from app.tasks.worker import verificar_compliance_drive
        verificar_compliance_drive.delay(str(report.id), nome_gerado)
    except Exception as e:
        # Não bloqueia a finalização se o Celery/Redis não estiver disponível
        import logging
        logging.getLogger(__name__).warning(
            f"Não foi possível disparar verificação de compliance: {e}"
        )

    return report


@router.post("/force-sync-drive", status_code=status.HTTP_202_ACCEPTED)
def forcar_sincronizacao_drive(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Dispara varredura geral para forçar a sincronização de relatórios Pendentes.
    """
    from app.tasks.worker import sincronizacao_noturna_drive
    try:
        sincronizacao_noturna_drive.delay()
        return {"message": "Sincronização em massa enviada para a fila com sucesso."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao enfileirar task de sync: {str(e)}")


@router.patch("/{report_id}/cancelar", response_model=ClassReportResponse)
def cancelar_relatorio(
    report_id: UUID,
    cancel_data: ClassReportCancel,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Cancela um relatório com justificativa obrigatória."""
    report = db.query(ClassReport).filter(ClassReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado.")

    if report.status == "CANCELADO":
        raise HTTPException(status_code=400, detail="Relatório já está cancelado.")

    # Restrição RBAC (Ownership)
    if current_user.role == "assistente" and report.created_by != current_user.id:
        raise HTTPException(
            status_code=403,
            detail="Assistentes só podem cancelar relatórios criados por eles mesmos.",
        )

    report.status = "CANCELADO"
    report.observacoes = (
        f"CANCELADO por {current_user.full_name}: {cancel_data.justificativa}"
    )

    db.commit()
    db.refresh(report)

    return report


@router.get("/", response_model=PaginatedReportsResponse)
def listar_relatorios(
    data_aula: Optional[date] = Query(None),
    data_inicio: Optional[date] = Query(None, description="Filtro período início"),
    data_fim: Optional[date] = Query(None, description="Filtro período fim"),
    turma_id: Optional[UUID] = Query(None),
    professor_id: Optional[UUID] = Query(None),
    disciplina_id: Optional[UUID] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    estudio: Optional[str] = Query(None),
    tipo_aula: Optional[str] = Query(None),
    com_atraso: Optional[bool] = Query(None, description="Filtrar com atraso"),
    com_substituicao: Optional[bool] = Query(None, description="Filtrar com substituição"),
    com_problema: Optional[bool] = Query(None, description="Filtrar com problema técnico"),
    sort_by: str = Query("data_aula", description="Campo de ordenação: data_aula ou created_at"),
    sort_order: str = Query("desc", description="Ordem: asc ou desc"),
    limit: int = Query(50, ge=1, le=200, description="Itens por página"),
    offset: int = Query(0, ge=0, description="Offset para paginação"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Lista relatórios com paginação server-side (§5.3).
    Padrão: 50 por página, ordenação decrescente por data_aula.
    """
    query = db.query(ClassReport)

    # Aplicar filtros
    query = _apply_report_filters(
        query,
        data_aula=data_aula,
        data_inicio=data_inicio,
        data_fim=data_fim,
        turma_id=turma_id,
        professor_id=professor_id,
        disciplina_id=disciplina_id,
        status_filter=status_filter,
        estudio=estudio,
        tipo_aula=tipo_aula,
        com_atraso=com_atraso,
        com_substituicao=com_substituicao,
        com_problema=com_problema,
    )

    # Contagem total (antes da paginação)
    total = query.count()

    # Ordenação
    sort_column = ClassReport.created_at if sort_by == "created_at" else ClassReport.data_aula
    if sort_order == "asc":
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())

    # Paginação
    items = query.offset(offset).limit(limit).all()

    return PaginatedReportsResponse(
        items=items,
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/export")
def exportar_relatorios(
    data_aula: Optional[date] = Query(None),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    turma_id: Optional[UUID] = Query(None),
    professor_id: Optional[UUID] = Query(None),
    disciplina_id: Optional[UUID] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    estudio: Optional[str] = Query(None),
    tipo_aula: Optional[str] = Query(None),
    com_atraso: Optional[bool] = Query(None),
    com_substituicao: Optional[bool] = Query(None),
    com_problema: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Exporta relatórios em XLSX (Excel) respeitando filtros ativos (§5.3).
    Formato idêntico ao modelo de saída, com nomes legíveis e links do Drive.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(ClassReport)

    query = _apply_report_filters(
        query,
        data_aula=data_aula,
        data_inicio=data_inicio,
        data_fim=data_fim,
        turma_id=turma_id,
        professor_id=professor_id,
        disciplina_id=disciplina_id,
        status_filter=status_filter,
        estudio=estudio,
        tipo_aula=tipo_aula,
        com_atraso=com_atraso,
        com_substituicao=com_substituicao,
        com_problema=com_problema,
    )

    reports = query.order_by(ClassReport.data_aula.desc()).all()

    # --- Pré-carregar nomes para evitar N+1 queries ---
    turma_ids = {r.turma_id for r in reports}
    disc_ids = {r.disciplina_id for r in reports}
    prof_ids = {r.professor_id for r in reports}
    sub_ids = {r.professor_substituto_id for r in reports if r.professor_substituto_id}

    turmas_map = {
        t.id: t for t in db.query(Turma).filter(Turma.id.in_(turma_ids)).all()
    } if turma_ids else {}
    discs_map = {
        d.id: d for d in db.query(Disciplina).filter(Disciplina.id.in_(disc_ids)).all()
    } if disc_ids else {}
    all_prof_ids = prof_ids | sub_ids
    profs_map = {
        p.id: p for p in db.query(Professor).filter(Professor.id.in_(all_prof_ids)).all()
    } if all_prof_ids else {}

    # Dias da semana em português
    DIAS_PT = {
        0: "Segunda", 1: "Terça", 2: "Quarta",
        3: "Quinta", 4: "Sexta", 5: "Sábado", 6: "Domingo",
    }

    # --- Criar workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Sistema"

    # Headers idênticos ao modelo
    headers = [
        "DATA", "DIA ", "TURNO", "ESTÚDIO", "CANAL",
        "CURSO", "HORÁRIO", "DISCIPLINA", "PROFESSOR",
        "REGULAR", "SUBSTITUIÇÃO DE PROFESSOR", "PROFESSOR QUE SUBSTITUIU",
        "INTERAÇÃO PROFESSOR ALUNO", "TIPO DE INTERAÇÃO",
        "ATIVIDADE PRÁTICA PARA ESCOLA", "TIPO DE ATIVIDADE PRÁTICA",
        "ATRASO PARA INÍCIO", "TEMPO DE ATRASO", "OBSERVAÇÃO ATRASO",
        "PROBLEMA COM MATERIAL ", "MOTIVO DO PROBLEMA",
        "UTILIZOU ALGUM RECURSO", "TIPO DE RECURSO", "FICHEIRO GERADO",
        "LINK DO VÍDEO", "PASTA DO DRIVE",
    ]

    # Estilo do header
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- Preencher dados ---
    for row_idx, r in enumerate(reports, 2):
        turma = turmas_map.get(r.turma_id)
        disc = discs_map.get(r.disciplina_id)
        prof = profs_map.get(r.professor_id)
        prof_sub = profs_map.get(r.professor_substituto_id) if r.professor_substituto_id else None

        # Calcular dia da semana
        dia_semana = DIAS_PT.get(r.data_aula.weekday(), "") if r.data_aula else ""

        # Data formatada DD/MM/YYYY
        data_fmt = r.data_aula.strftime("%d/%m/%Y") if r.data_aula else ""

        # Interação: campo pode ser "Não", "Chat", "Videoconferência", etc.
        tem_interacao = r.interacao_professor_aluno and r.interacao_professor_aluno != "Não"
        interacao_sim_nao = "Sim" if tem_interacao else "Não"
        tipo_interacao = r.interacao_professor_aluno if tem_interacao else ""

        # Atividade prática
        tem_atividade = bool(r.atividade_pratica and r.atividade_pratica.strip())
        atividade_sim_nao = "Sim" if tem_atividade else "Não"

        # Atraso
        atraso_sim_nao = "sim" if r.teve_atraso else "não"
        tempo_atraso = str(r.minutos_atraso) if r.teve_atraso and r.minutos_atraso else "não"

        # Problema material
        tem_problema = r.problema_material and r.problema_material != "Não"
        problema_sim_nao = "Sim" if tem_problema else "Não"
        motivo_problema = r.problema_material if tem_problema else "Não"

        # Recursos
        tem_recurso = bool(r.tipo_recursos_utilizados and len(r.tipo_recursos_utilizados) > 0)
        recurso_sim_nao = "Sim" if tem_recurso else "Não"
        tipo_recurso = ", ".join(r.tipo_recursos_utilizados) if tem_recurso else ""

        # Drive links
        video_link = r.video_link or ""
        try:
            folder_link = r.video_folder_link or ""
        except Exception:
            folder_link = ""

        row_data = [
            data_fmt,
            dia_semana,
            r.turno or "",
            r.estudio or "",
            r.canal_utilizado or "",
            turma.nome if turma else "",
            r.horario_aula or "",
            disc.nome if disc else "",
            prof.nome if prof else "",
            r.regular or "Sim",
            "Sim" if r.teve_substituicao else "Não",
            prof_sub.nome if prof_sub else "-",
            interacao_sim_nao,
            tipo_interacao,
            atividade_sim_nao,
            r.atividade_pratica or "",
            atraso_sim_nao,
            tempo_atraso,
            r.observacao_atraso or "",
            problema_sim_nao,
            motivo_problema,
            recurso_sim_nao,
            tipo_recurso,
            r.nome_ficheiro_gerado or "",
            video_link,
            folder_link,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Ajustar largura das colunas
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(len(reports) + 2, 50))
        ) if reports else len(headers[col_idx - 1])
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    # --- Gerar resposta ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"relatorios_{date.today().strftime('%d.%m.%y')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


@router.get("/{report_id}", response_model=ClassReportResponse)
def obter_relatorio(
    report_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Obtém detalhes de um relatório específico."""
    report = db.query(ClassReport).filter(ClassReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado.")
    return report


@router.put("/{report_id}", response_model=ClassReportResponse)
def editar_relatorio_admin(
    report_id: UUID,
    data: ClassReportCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """Admin: Edita qualquer relatório e salva no log."""
    report = db.query(ClassReport).filter(ClassReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado.")

    _validar_professor_disciplina(db, data.professor_id, data.disciplina_id)
    _validar_campos_condicionais(data)

    for key, value in data.model_dump().items():
        setattr(report, key, value)

    if report.status == "FINALIZADO":
        turma = db.query(Turma).filter(Turma.id == report.turma_id).first()
        disciplina = db.query(Disciplina).filter(Disciplina.id == report.disciplina_id).first()
        nome_gerado = gerar_nome_padronizado(
            nomenclatura_turma=turma.nomenclatura_padrao or turma.nome,
            disciplina=disciplina.nome if disciplina else "",
            data_aula=report.data_aula,
            conteudo=report.conteudo_ministrado,
        )
        report.nome_ficheiro_gerado = nome_gerado

    log = AuditLog(
        actor_id=current_user.id,
        action="EDIT_REPORT",
        details=f"Admin editou o relatório {report.id}"
    )
    db.add(log)
    db.commit()
    db.refresh(report)
    return report


@router.delete("/{report_id}", status_code=status.HTTP_204_NO_CONTENT)
def excluir_relatorio_admin(
    report_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    """Admin: Exclui um relatório permanentemente e salva no log."""
    report = db.query(ClassReport).filter(ClassReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Relatório não encontrado.")

    log = AuditLog(
        actor_id=current_user.id,
        action="DELETE_REPORT",
        details=f"Admin excluiu o relatório {report.id} (Turma: {report.turma_id}, Data: {report.data_aula})"
    )
    db.add(log)
    db.delete(report)
    db.commit()
    return None


# ─── Endpoints Auxiliares (Lookup Data) ──────────────────────────────


@router.get("/lookup/professores-por-disciplina/{disciplina_id}")
def listar_professores_por_disciplina(
    disciplina_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Retorna apenas os professores vinculados a uma disciplina específica.
    Usado pelo frontend para preencher o dropdown de professores.
    """
    professores = (
        db.query(Professor)
        .join(ProfessorDisciplina)
        .filter(ProfessorDisciplina.disciplina_id == disciplina_id)
        .all()
    )
    return [{"id": str(p.id), "nome": p.nome} for p in professores]


@router.get("/lookup/disciplinas")
def listar_disciplinas(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Retorna todas as disciplinas cadastradas."""
    disciplinas = db.query(Disciplina).order_by(Disciplina.nome).all()
    return [{"id": str(d.id), "nome": d.nome} for d in disciplinas]


@router.get("/lookup/professores")
def listar_professores(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Retorna todos os professores cadastrados."""
    professores = db.query(Professor).order_by(Professor.nome).all()
    return [{"id": str(p.id), "nome": p.nome} for p in professores]
